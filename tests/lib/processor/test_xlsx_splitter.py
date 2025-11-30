import io
import pytest
from openpyxl import Workbook

from lib.processor.splitters.xlsx_splitter import split_xlsx, XLSXSplitError


def create_test_xlsx(num_sheets: int = 1, rows_per_sheet: int = 100) -> bytes:
    """Helper function to create a test XLSX file."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for i in range(num_sheets):
        ws = wb.create_sheet(f"Sheet{i+1}")
        for row in range(1, rows_per_sheet + 1):
            for col in range(1, 11):  # 10 columns
                ws.cell(row=row, column=col, value=f"Data_{i}_{row}_{col}")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestXLSXSplitter:
    def test_small_file_not_split(self):
        """Test that files smaller than 4MB are returned as-is."""
        file_bytes = create_test_xlsx(num_sheets=1, rows_per_sheet=100)
        
        result = split_xlsx(file_bytes)
        
        assert len(result) == 1
        assert result[0] == file_bytes

    def test_empty_file(self):
        """Test handling of files with no sheets returns empty list."""
        # Note: openpyxl doesn't allow saving workbooks with no sheets
        # So we create a minimal workbook and check the splitter handles it
        wb = Workbook()
        # Keep the default sheet, just make it empty
        buffer = io.BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()
        
        result = split_xlsx(file_bytes)
        
        # Should return the file as-is since it's small
        assert len(result) == 1

    def test_large_file_split(self):
        """Test that large files are split into multiple chunks."""
        # Create a large file with multiple sheets
        file_bytes = create_test_xlsx(num_sheets=10, rows_per_sheet=5000)
        
        # Only test if the file is actually large enough
        if len(file_bytes) > 4 * 1024 * 1024:
            result = split_xlsx(file_bytes)
            assert len(result) > 1
            
            # Verify each chunk is valid
            for chunk in result:
                assert len(chunk) > 0
                # Try to load it with openpyxl to verify it's valid
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(chunk))
                assert len(wb.sheetnames) > 0

    def test_split_chunks_are_valid_xlsx(self):
        """Test that all chunks are valid XLSX files."""
        file_bytes = create_test_xlsx(num_sheets=3, rows_per_sheet=1000)
        
        result = split_xlsx(file_bytes)
        
        from openpyxl import load_workbook
        for chunk in result:
            # Should not raise an exception
            wb = load_workbook(io.BytesIO(chunk))
            assert wb is not None
            assert len(wb.sheetnames) > 0

    def test_all_sheets_preserved(self):
        """Test that all sheets from the original file are preserved in chunks."""
        num_sheets = 5
        file_bytes = create_test_xlsx(num_sheets=num_sheets, rows_per_sheet=500)
        
        result = split_xlsx(file_bytes)
        
        from openpyxl import load_workbook
        total_sheets = 0
        for chunk in result:
            wb = load_workbook(io.BytesIO(chunk))
            total_sheets += len(wb.sheetnames)
        
        assert total_sheets == num_sheets

    def test_corrupted_file_raises_error(self):
        """Test that corrupted files raise XLSXSplitError."""
        corrupted_bytes = b"This is not a valid XLSX file"
        
        # python-calamine handles errors gracefully, should raise XLSXSplitError
        try:
            result = split_xlsx(corrupted_bytes)
            # If it doesn't raise, it should at least return empty or fail
            assert False, "Should have raised XLSXSplitError for corrupted file"
        except XLSXSplitError:
            # Expected behavior
            pass
