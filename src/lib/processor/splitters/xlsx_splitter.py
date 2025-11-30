import io
import logging
from typing import List

from openpyxl import Workbook, load_workbook
from python_calamine import CalamineWorkbook

logger = logging.getLogger(__name__)


class XLSXSplitError(Exception):
    pass


def split_xlsx(file_bytes: bytes) -> List[bytes]:
    """
    Splits a binary XLSX blob into a list of valid XLSX blobs,
    where each one has an approximate size < 4MB.

    The split is done by complete sheets, preserving the validity
    of the Excel structure.
    """
    TARGET_LIMIT_BYTES = 4 * 1024 * 1024  # 4 MB hard limit
    # We use a "soft" target to leave room for headers and metadata
    SOFT_TARGET_BYTES = int(TARGET_LIMIT_BYTES * 0.90)

    chunks = []

    try:
        # Validate that the file is a valid XLSX before processing
        # This ensures that corrupted files throw an error even if they are small
        workbook_calamine = CalamineWorkbook.from_filelike(io.BytesIO(file_bytes))
        sheet_names = workbook_calamine.sheet_names

        if not sheet_names:
            return []

        # Fast-path: Si ya cabe, devolvemos tal cual
        if len(file_bytes) < TARGET_LIMIT_BYTES:
            return [file_bytes]

        # Cargar el workbook con openpyxl para poder escribir chunks
        wb = load_workbook(io.BytesIO(file_bytes))

        current_sheets = []
        current_size_estimate = 0

        for sheet_name in sheet_names:
            # Create a temporary workbook to estimate the size of this sheet
            temp_wb = Workbook()
            temp_wb.remove(temp_wb.active)  # Remove the default sheet

            # Copy the current sheet
            source_sheet = wb[sheet_name]
            target_sheet = temp_wb.create_sheet(sheet_name)

            # Copy cells (simplified - copies values and formulas)
            for row in source_sheet.iter_rows():
                for cell in row:
                    target_cell = target_sheet[cell.coordinate]
                    target_cell.value = cell.value
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy()
                        target_cell.alignment = cell.alignment.copy()

            # Save to buffer to estimate size
            buffer = io.BytesIO()
            temp_wb.save(buffer)
            sheet_size = len(buffer.getvalue())

            # CASE 1: If adding this sheet exceeds the limit and we already have accumulated sheets
            if current_size_estimate + sheet_size > SOFT_TARGET_BYTES and current_sheets:
                # Save the current chunk
                chunk_wb = Workbook()
                chunk_wb.remove(chunk_wb.active)

                for sheet_to_copy in current_sheets:
                    source = wb[sheet_to_copy]
                    target = chunk_wb.create_sheet(sheet_to_copy)

                    for row in source.iter_rows():
                        for cell in row:
                            target_cell = target[cell.coordinate]
                            target_cell.value = cell.value
                            if cell.has_style:
                                target_cell.font = cell.font.copy()
                                target_cell.border = cell.border.copy()
                                target_cell.fill = cell.fill.copy()
                                target_cell.number_format = cell.number_format
                                target_cell.protection = cell.protection.copy()
                                target_cell.alignment = cell.alignment.copy()

                chunk_buffer = io.BytesIO()
                chunk_wb.save(chunk_buffer)
                chunk_data = chunk_buffer.getvalue()

                if len(chunk_data) > TARGET_LIMIT_BYTES:
                    logger.warning(
                        f"Chunk size ({len(chunk_data)} bytes) exceeds target limit. "
                        f"Sheet(s) {current_sheets} are too large individually."
                    )

                chunks.append(chunk_data)

                # Reset for the next sheet
                current_sheets = [sheet_name]
                current_size_estimate = sheet_size

            # CASE 2: Add this sheet to the current chunk
            else:
                current_sheets.append(sheet_name)
                current_size_estimate += sheet_size

        # Save the last chunk if there are pending sheets
        if current_sheets:
            chunk_wb = Workbook()
            chunk_wb.remove(chunk_wb.active)

            for sheet_to_copy in current_sheets:
                source = wb[sheet_to_copy]
                target = chunk_wb.create_sheet(sheet_to_copy)

                for row in source.iter_rows():
                    for cell in row:
                        target_cell = target[cell.coordinate]
                        target_cell.value = cell.value
                        if cell.has_style:
                            target_cell.font = cell.font.copy()
                            target_cell.border = cell.border.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.number_format = cell.number_format
                            target_cell.protection = cell.protection.copy()
                            target_cell.alignment = cell.alignment.copy()

            chunk_buffer = io.BytesIO()
            chunk_wb.save(chunk_buffer)
            chunks.append(chunk_buffer.getvalue())

    except Exception as e:
        logger.exception(e)
        raise XLSXSplitError(f"Critical error processing XLSX: {e}") from e

    return chunks
